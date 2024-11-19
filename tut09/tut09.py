import pandas as pd
data = pd.read_csv('input_attendance.csv')

timestamps_set = data['Timestamp'].unique()

print("First 10 Timestamp Values:", timestamps_set[:10])

data['Timestamp'] = pd.to_datetime(data['Timestamp'], errors='coerce', dayfirst=True, infer_datetime_format=True)

invalid_entries = data[data['Timestamp'].isna()]

print(f"Number of invalid timestamps after processing: {len(invalid_entries)}")

print(invalid_entries[['Timestamp', 'Roll']].head(10))

accepted_dates = pd.to_datetime([
    "06/08/2024", "13/08/2024", "20/08/2024", "27/08/2024",
    "03/09/2024", "17/09/2024", "01/10/2024"
], dayfirst=True)

roll_attendance_list = []

for roll in data['Roll'].unique():
    count_per_date = {'Roll': roll}
    for date in accepted_dates:
        # Filter records where both roll number and timestamp date match
        count_per_date[date.strftime('%d/%m/%Y')] = data[
            (data['Roll'] == roll) &
            (data['Timestamp'].dt.date == date.date())
        ].shape[0]
    roll_attendance_list.append(count_per_date)

attendance_summary = pd.DataFrame(roll_attendance_list)

attendance_summary['Total Marked'] = attendance_summary.iloc[:, 1:].sum(axis=1)

total_entries = data.groupby('Roll').size().reset_index(name='Overall Attendance')

final_summary = attendance_summary.merge(total_entries, on='Roll', how='left')

final_summary['Inconsistencies'] = (final_summary['Overall Attendance'] - final_summary['Total Marked']).abs()

output_file_name = 'final_attendance_summary.csv'
final_summary.to_csv(output_file_name, index=False)

import pandas as pd

final_summary = pd.read_csv('final_attendance_summary.csv')

with open('stud_list.txt', 'r') as file:
    student_data = file.readlines()

student_data = [entry.strip() for entry in student_data]

students = pd.DataFrame(student_data, columns=['FullEntry'])

combined_attendance = final_summary.copy()

for record in students['FullEntry']:
    # If the student is not already listed in the attendance DataFrame
    if record not in final_summary['Roll'].astype(str).values:
        # Split the entry into roll number and name
        roll_num, name = record.split(' ', 1)
        # Create a row with the roll number and default attendance counts
        new_entry = {'Roll': roll_num}
        for col in final_summary.columns[1:]:
            new_entry[col] = 0
        combined_attendance = pd.concat([combined_attendance, pd.DataFrame([new_entry])], ignore_index=True)

new_output_file = 'combined_attendance_summary.csv'
combined_attendance.to_csv(new_output_file, index=False)

new_output_file

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

summary_data = pd.read_csv('combined_attendance_summary.csv')

output_xlsx_file = 'highlighted_attendance_summary.xlsx'
wb = openpyxl.Workbook()
sheet = wb.active

valid_attendance_dates = [
    pd.Timestamp('2024-08-06'), pd.Timestamp('2024-08-13'),
    pd.Timestamp('2024-08-20'), pd.Timestamp('2024-08-27'),
    pd.Timestamp('2024-09-03'), pd.Timestamp('2024-09-17'),
    pd.Timestamp('2024-10-01')
]

sheet.append(['Roll'] + [date.strftime('%d/%m/%Y') for date in valid_attendance_dates])

for idx, entry in summary_data.iterrows():

    sheet.append(entry.values.tolist())

    for col_idx, count in enumerate(entry[1:], start=2):  # Start from second column (B)
        if count == 0:
            fill_color = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
        elif count == 1:
            fill_color = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
        elif count == 2:
            fill_color = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green
        else:
            fill_color = PatternFill(fill_type=None)  # No fill

        sheet.cell(row=idx + 2, column=col_idx).fill = fill_color  # Adjust for row offset

wb.save(output_xlsx_file)